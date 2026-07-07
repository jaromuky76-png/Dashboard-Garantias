import os
import json
import datetime
import openpyxl
import urllib.request
import re
import time

# Configuración API Gestioo y caché
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_PATH = os.path.join(BASE_DIR, "api_cache.json")
BASE_API_URL = "https://taller.gestioo.net/taller/consulta/obtener_orden/"
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

SKIP_KEYWORDS = [
    "DIAGNOSTICO", "MANTENIMIENTO", "MANO DE OBRA", "REVISION", 
    "LIMPIEZA", "AJUSTE", "CALIBRACION", "TRANSPORTE", "FLETE",
    "VISITA", "LEVANTAMIENTO", "ACTIVIDAD TECNICA", "INSTALACION", 
    "ARMADO", "CHEQUEO", "REPARACION", "PRUEBA", "DESPLAZAMIENTO",
    "SERVICIOS", "MANO OBRA"
]

MESES_MAP = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
    'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
}

output_js  = os.path.join(BASE_DIR, "data.js")
output_svc = os.path.join(BASE_DIR, "servicio_data.js")
output_pts = os.path.join(BASE_DIR, "parts_data.js")
output_seg = os.path.join(BASE_DIR, "seguimiento_data.js")

def load_cache():
    if os.path.exists(CACHE_PATH):
        try:
            with open(CACHE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            pass
    return {}

def save_cache(cache):
    try:
        with open(CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        pass

def clean_text(text):
    if text is None: return ""
    text = str(text).upper().strip()
    import unicodedata
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    return text

def is_spare_part(description):
    desc = clean_text(description)
    for kw in SKIP_KEYWORDS:
        if kw in desc:
            return False
    return True

def extract_order_code(url):
    if not isinstance(url, str): return None
    match = re.search(r'/orden/([^/?#]+)', url)
    if match: return match.group(1)
    return None

def extract_part_code(item):
    code = item.get("codigo")
    if code:
        code_str = str(code).strip()
        if len(code_str) == 9 and code_str.isdigit():
            return code_str
    desc = str(item.get("descripcion") or "")
    if desc:
        match = re.search(r'\b(\d{9})\b', desc.strip())
        if match: return match.group(1)
        match_start = re.match(r'^(\d{9})', desc.strip())
        if match_start: return match_start.group(1)
    return "N/A"

def clean_item_description(desc, part_code):
    if not desc: return ""
    desc_str = str(desc).strip()
    if part_code != "N/A" and desc_str.startswith(part_code):
        desc_str = desc_str[len(part_code):].strip(" -:.")
    return desc_str

def fetch_order_parts(order_code, cache):
    if order_code in cache: return cache[order_code]
    url = f"{BASE_API_URL}{order_code}"
    print(f"  -> Consultando API para {order_code}...")
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))
            if data.get('error') == 0:
                parts = data.get('datos', {}).get('servicios_repuestos', [])
                cache[order_code] = parts
                save_cache(cache)
                time.sleep(1)
                return parts
    except Exception as e:
        print(f"     Error al consultar {order_code}: {e}")
    return []

def safe_str(v, default=''):
    if v is None: return default
    return str(v).strip()

def safe_int(v, default=1):
    try:
        return int(v) if v is not None else default
    except (ValueError, TypeError):
        return default

def fmt_datetime(v):
    if isinstance(v, datetime.datetime): return v.strftime('%Y-%m-%d %H:%M')
    if isinstance(v, datetime.date): return v.strftime('%Y-%m-%d 00:00')
    return ''

def load_existing_json(filepath, var_name):
    if not os.path.exists(filepath): return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            match = re.search(r'' + var_name + r'\s*=\s*(\[.*\]);', content, flags=re.DOTALL)
            if match: return json.loads(match.group(1))
    except Exception as e:
        print(f"Error cargando {filepath}: {e}")
    return []

def save_json(filepath, var_name, data, desc, meta_var_name):
    try:
        archivos = set()
        for row in data:
            u = row.get('unidad_negocio', '')
            a = row.get('anio', '')
            m = row.get('mes', '')
            archivos.add(f"{u}-{a}-{m}")
            
        meta = {
            "totalRegistros": len(data),
            "archivosProcessados": len(archivos),
            "errores": 0
        }
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"// Generado por dashboard_updater.py -- {desc}\n")
            f.write(f"window.{var_name} = {json.dumps(data, ensure_ascii=False)};\n")
            f.write(f"window.{meta_var_name} = {json.dumps(meta, ensure_ascii=False)};\n")
    except Exception as e:
        print(f"Error guardando {filepath}: {e}")

def remove_existing_period(data, unidad, anio, mes):
    return [x for x in data if not (str(x.get('anio')) == str(anio) and str(x.get('mes')).upper() == mes.upper() and str(x.get('unidad_negocio')).upper() == unidad.upper())]

def process_single_file(filepath, unidad, anio, mes, mes_num):
    print(f"Iniciando actualización incremental: {unidad} {mes} {anio}...")
    api_cache = load_cache()
    
    garantia_data = load_existing_json(output_js, 'PRELOADED_DATA')
    servicio_data = load_existing_json(output_svc, 'PRELOADED_SERVICIO')
    parts_data = load_existing_json(output_pts, 'partsData')
    seguimiento_data = load_existing_json(output_seg, 'PRELOADED_SEGUIMIENTO')
    
    # Remover registros viejos (excepto para seguimiento, que mantiene el historial)
    garantia_data = remove_existing_period(garantia_data, unidad, anio, mes)
    servicio_data = remove_existing_period(servicio_data, unidad, anio, mes)
    parts_data = remove_existing_period(parts_data, unidad, anio, mes)
    
    cnt_g, cnt_s, cnt_p, cnt_seg = 0, 0, 0, 0
    existing_seg = {f"{x.get('unidad_negocio', '')}-{x.get('ot', '')}" for x in seguimiento_data}
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ot_sheet_name = None
        for name in wb.sheetnames:
            if name.strip().upper() == 'OT' or name.strip().upper() == 'ESTADO DE OT':
                ot_sheet_name = name
                break
        if not ot_sheet_name:
            ot_sheet_name = wb.sheetnames[0]
            
        ot_sheet = wb[ot_sheet_name]
        ot_iterator = ot_sheet.iter_rows(values_only=True)
        ot_headers = None
        
        for row in ot_iterator:
            row_list = list(row)
            if any(v and (str(v).strip().upper() == 'MARCA' or 'TIPO DE GARANTIA' in str(v).upper()) for v in row_list):
                ot_headers = row_list
                break
                
        if ot_headers:
            h = {str(v).strip().upper(): i for i, v in enumerate(ot_headers) if v}
            # Unificar indices
            ot_i = h.get('NO. OT', h.get('NO. OT/MR', -1))
            fecha_i = h.get('FECHA Y HR DE INGRESO', -1)
            marca_i = h.get('MARCA', -1)
            rms_i = h.get('RMS', h.get('RMS EVAPORADOR', -1))
            desc_i = h.get('DESCRIPCION DEL EQUIPO', h.get('DESCRIPCION DEL EVAPORADOR', -1))
            cant_i = h.get('CANT2', h.get('CANT', -1))
            tipo_i = h.get('TIPO DE GARANTIA', -1)
            acep_i = h.get('ACEPTACION', -1)
            link_i = h.get('LINK OT DIGITAL', -1)
            
            for row in ot_iterator:
                row = list(row)
                
                tipo = safe_str(row[tipo_i] if tipo_i >= 0 else None).upper()
                if not tipo: continue
                
                acep = safe_str(row[acep_i] if acep_i >= 0 else None).upper()
                marca = safe_str(row[marca_i] if marca_i >= 0 else None, 'DESCONOCIDA').upper()
                rms = safe_str(row[rms_i] if rms_i >= 0 else None, 'N/A')
                desc = safe_str(row[desc_i] if desc_i >= 0 else None, '')
                cant = safe_int(row[cant_i] if cant_i >= 0 else None, 1)
                ot_n = safe_str(row[ot_i] if ot_i >= 0 else None, '')
                link = safe_str(row[link_i] if link_i >= 0 else None, '')
                fecha = fmt_datetime(row[fecha_i] if fecha_i >= 0 else None)
                
                base = {
                    "marca": marca, "rms": rms, "descripcion": desc,
                    "fecha": fecha, "anio": int(anio), "mes": mes, "mesNum": int(mes_num),
                    "ot": ot_n, "unidad_negocio": unidad.upper()
                }
                
                # Garantias y Servicios
                if "GARANTIA" in tipo:
                    norm = "GARANTIA TOTAL" if "TOTAL" in tipo else "GARANTIA PARCIAL"
                    garantia_data.append({**base, "tipoGarantia": norm})
                    cnt_g += 1
                    
                    # Agregar al seguimiento si no existe
                    if f"{unidad.upper()}-{ot_n}" not in existing_seg and ot_n:
                        seguimiento_data.append({
                            "ot": ot_n, "marca": marca, "descripcion": desc,
                            "fecha_ingreso": fecha, "unidad_negocio": unidad.upper(),
                            "anio": int(anio), "mes": mes, "mesNum": int(mes_num),
                            "estado": "Pendiente", "notas": "", "fecha_estimada": ""
                        })
                        existing_seg.add(f"{unidad.upper()}-{ot_n}")
                        cnt_seg += 1
                elif "SERVICIO" in tipo:
                    servicio_data.append(base)
                    cnt_s += 1
                    
                # Repuestos
                es_parcial = 'PARCIAL' in tipo
                es_svc_apro = 'SERVICIO' in tipo and acep == 'SI'
                if es_parcial or es_svc_apro:
                    order_code = extract_order_code(link)
                    if order_code:
                        parts = fetch_order_parts(order_code, api_cache)
                        for p in parts:
                            desc_part = p.get('descripcion', '')
                            if is_spare_part(desc_part):
                                part_code = extract_part_code(p)
                                clean_desc = clean_item_description(desc_part, part_code)
                                cant_part = safe_int(p.get('cantidad'), 1)
                                parts_data.append({
                                    "ot": ot_n, "marca": marca, "codigo_repuesto": part_code,
                                    "descripcion": clean_desc, "cantidad": cant_part,
                                    "tipo_garantia": tipo, "aceptacion": acep, "link": link,
                                    "fecha": fecha, "anio": int(anio), "mes": mes, "mesNum": int(mes_num),
                                    "unidad_negocio": unidad.upper()
                                })
                                cnt_p += 1
    except Exception as e:
        print(f"Error procesando {filepath}: {e}")
        import traceback; traceback.print_exc()
    finally:
        try:
            if 'wb' in locals() and wb:
                wb.close()
        except:
            pass

    # Guardar los 4 archivos
    save_json(output_js, 'PRELOADED_DATA', garantia_data, "Garantias", "PRELOADED_META")
    save_json(output_svc, 'PRELOADED_SERVICIO', servicio_data, "Servicios", "PRELOADED_META_SVC")
    save_json(output_pts, 'partsData', parts_data, "Repuestos", "PARTS_META")
    save_json(output_seg, 'PRELOADED_SEGUIMIENTO', seguimiento_data, "Seguimiento", "SEGUIMIENTO_META")
    
    print(f"Finalizado: {cnt_g} garantias, {cnt_s} servicios, {cnt_p} repuestos, {cnt_seg} nuevos seguimientos agregados.")
    return True
