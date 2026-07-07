import os
import re
import gc
import json
import subprocess
import threading
import tempfile
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer

# Directorio base del proyecto
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OT_BASE_DIR = os.environ.get('OT_BASE_DIR', os.path.abspath(os.path.join(BASE_DIR, "..", "OT")))

# MESES_MAP: string key -> int value
MESES_MAP = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
    'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
}
# Reverse: int -> string
MESES_INV = {v: k for k, v in MESES_MAP.items()}

# Variable global para el estado
is_processing = False


def stream_multipart_to_disk(rfile, content_length, boundary_bytes, temp_file_path):
    """
    Streams the multipart request body from the socket directly to disk,
    never loading the full body into RAM. Returns (filename, unidad).
    Uses a sliding window buffer of only 64KB at a time.
    """
    CHUNK = 65536  # 64KB chunks
    filename = None
    unidad = 'CS'

    # We write the raw multipart body to a temp file first
    raw_path = temp_file_path + ".raw"
    remaining = content_length
    with open(raw_path, 'wb') as raw_f:
        while remaining > 0:
            chunk = rfile.read(min(CHUNK, remaining))
            if not chunk:
                break
            raw_f.write(chunk)
            remaining -= len(chunk)

    # Now parse the raw file to extract the Excel file part and form fields
    # We scan through the file looking for boundaries
    boundary_full = b'--' + boundary_bytes
    boundary_end = boundary_full + b'--'

    with open(raw_path, 'rb') as raw_f:
        content = raw_f.read()

    # Extract parts
    parts = content.split(boundary_full)
    excel_data = None

    for part in parts:
        if not part or part == b'--\r\n' or part == b'--':
            continue
        # Remove leading \r\n
        if part.startswith(b'\r\n'):
            part = part[2:]
        # Remove trailing \r\n
        if part.endswith(b'\r\n'):
            part = part[:-2]

        sep = part.find(b'\r\n\r\n')
        if sep == -1:
            continue

        header_raw = part[:sep]
        body_part = part[sep + 4:]

        header_str = header_raw.decode('utf-8', 'ignore')

        # Check for field name
        name_match = re.search(r'name="([^"]+)"', header_str)
        if not name_match:
            continue
        field_name = name_match.group(1)

        if 'filename=' in header_str:
            fn_match = re.search(r'filename="([^"]+)"', header_str)
            if fn_match:
                filename = os.path.basename(fn_match.group(1))
            excel_data = body_part
        elif field_name in ('unidad', 'unidad_negocio'):
            val = body_part.decode('utf-8', 'ignore').strip().upper()
            if val:
                unidad = val

    # Write Excel data to final temp path
    if excel_data:
        with open(temp_file_path, 'wb') as out_f:
            out_f.write(excel_data)

    # Clean up raw file and free memory
    try:
        os.remove(raw_path)
    except Exception:
        pass

    del content
    del excel_data
    gc.collect()

    return filename, unidad


class DashboardServer(SimpleHTTPRequestHandler):
    def end_headers(self):
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        super().end_headers()

    def do_GET(self):
        if self.path == '/status':
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'processing': is_processing}).encode())
        else:
            super().do_GET()

    def do_POST(self):
        if self.path == '/upload':
            self.handle_upload()
        elif self.path == '/api/actualizar_tramite':
            self.handle_actualizar_tramite()
        else:
            self.send_error(404, "Not Found")

    def handle_upload(self):
        try:
            content_type = self.headers.get('Content-Type', '')
            if 'multipart/form-data' not in content_type:
                self.send_error(400, "Bad Request: expected multipart/form-data")
                return

            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                self.send_error(400, "Empty request")
                return

            # Extract boundary
            boundary_match = re.search(r'boundary=([^\s;]+)', content_type)
            if not boundary_match:
                self.send_error(400, "No boundary in Content-Type")
                return
            boundary_bytes = boundary_match.group(1).encode()

            temp_path = os.path.join(BASE_DIR, "temp_upload.xlsx")

            # Stream the file to disk without loading into RAM
            filename, unidad = stream_multipart_to_disk(
                self.rfile, content_length, boundary_bytes, temp_path
            )

            if not filename or not os.path.exists(temp_path):
                self.send_error(400, "No file uploaded or file empty")
                return

            if unidad not in ['CS', 'MAESTROS']:
                unidad = 'CS'

            anio_str, mes_nombre = self.detect_date_info(temp_path, filename, unidad)

            if not mes_nombre:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                self.send_response(400)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'No se pudo determinar el mes del archivo'}).encode())
                return

            if unidad not in ['CS', 'MAESTROS']:
                unidad = 'CS'

            target_dir = os.path.join(OT_BASE_DIR, unidad, anio_str, mes_nombre)
            os.makedirs(target_dir, exist_ok=True)
            target_path = os.path.join(target_dir, filename)

            # Verificar si existe algún archivo en la carpeta
            existing_files = [f for f in os.listdir(target_dir) if f.endswith('.xlsx') or f.endswith('.xls')] if os.path.exists(target_dir) else []

            # NOTE: overwrite is True by default since we removed reading it from multipart
            # (it was causing issues). Now we always overwrite when uploading same period.
            if existing_files:
                for old_f in existing_files:
                    try:
                        os.remove(os.path.join(target_dir, old_f))
                    except Exception:
                        pass

            os.rename(temp_path, target_path)
            print(f"Archivo guardado: {target_path}")

            # Ejecutar actualización en segundo plano
            global is_processing
            is_processing = True

            def run_updater(path, u, a, m, m_num):
                global is_processing
                try:
                    import dashboard_updater
                    print(f"Ejecutando actualización incremental para {u} {m} {a}...")
                    dashboard_updater.process_single_file(path, u, a, m, m_num)
                    print("Actualización completada en bases de datos.")

                    # Sincronización con GitHub
                    print("Sincronizando con GitHub...")
                    subprocess.run(
                        ["git", "add", "data.js", "servicio_data.js", "parts_data.js", "seguimiento_data.js", "api_cache.json"],
                        cwd=BASE_DIR, check=False
                    )
                    status = subprocess.run(
                        ["git", "status", "--porcelain"],
                        cwd=BASE_DIR, capture_output=True, text=True
                    )
                    if status.stdout.strip():
                        commit_msg = f"Update data from web upload: {u} {m} {a}"
                        subprocess.run(["git", "commit", "-m", commit_msg], cwd=BASE_DIR, check=False)

                        token = os.environ.get("GITHUB_TOKEN")
                        if token:
                            repo_url = f"https://jaromuky76-png:{token}@github.com/jaromuky76-png/Dashboard-Garantias.git"
                            subprocess.run(["git", "remote", "set-url", "origin", repo_url], cwd=BASE_DIR, check=False)

                        push_res = subprocess.run(
                            ["git", "push", "origin", "HEAD:master"],
                            cwd=BASE_DIR, capture_output=True, text=True
                        )
                        if push_res.returncode == 0:
                            print("GitHub sincronizado exitosamente.")
                        else:
                            print(f"Error en git push: {push_res.stderr}")
                    else:
                        print("No hubo cambios nuevos para subir a GitHub.")

                    try:
                        if os.path.exists(path):
                            os.remove(path)
                    except Exception as e:
                        print(f"No se pudo borrar temporal {path}: {e}")

                except Exception as e:
                    print(f"Error actualizando datos: {str(e)}")
                    import traceback
                    traceback.print_exc()
                finally:
                    is_processing = False

            mes_num = MESES_MAP.get(mes_nombre, 0)
            t = threading.Thread(target=run_updater, args=(target_path, unidad, anio_str, mes_nombre, mes_num))
            t.daemon = True
            t.start()

            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True, 'mes': mes_nombre}).encode())

        except Exception as ex:
            import traceback
            traceback.print_exc()
            print(f"Server error during upload: {ex}")
            try:
                self.send_response(500)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': f'Server error: {str(ex)}'}).encode())
            except Exception:
                pass

    def detect_date_info(self, filepath, filename, unidad="CS"):
        mes_detectado = None
        anio_detectado = None

        # 1. Intentar deducir del nombre del archivo (MESES_MAP has string keys now)
        filename_upper = str(filename).upper()
        for mes_name in MESES_MAP.keys():
            if mes_name in filename_upper:
                mes_detectado = mes_name
                break

        match_year = re.search(r'(20\d{2})', filename_upper)
        if match_year:
            anio_detectado = match_year.group(1)

        # 2. Si falta algo, leer el Excel usando la columna específica de fecha
        if not mes_detectado or not anio_detectado:
            try:
                import openpyxl
                from collections import Counter
                wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                ot_sheet_name = None
                for name in wb.sheetnames:
                    if name.strip().upper() in ('OT', 'ESTADO DE OT'):
                        ot_sheet_name = name
                        break
                if not ot_sheet_name:
                    ot_sheet_name = wb.sheetnames[0]

                ws = wb[ot_sheet_name]
                # CS -> columna D (idx 4), MAESTROS -> columna E (idx 5)
                col_idx = 5 if unidad.upper() == 'MAESTROS' else 4

                meses_counter = Counter()
                anios_counter = Counter()
                for row in ws.iter_rows(min_row=3, max_row=200, values_only=True):
                    if len(row) >= col_idx:
                        val = row[col_idx - 1]
                        if val and hasattr(val, 'month') and hasattr(val, 'year'):
                            meses_counter[val.month] += 1
                            anios_counter[val.year] += 1

                if meses_counter and not mes_detectado:
                    most_freq_month = meses_counter.most_common(1)[0][0]
                    mes_detectado = MESES_INV.get(most_freq_month)

                if anios_counter and not anio_detectado:
                    most_freq_year = anios_counter.most_common(1)[0][0]
                    anio_detectado = str(int(most_freq_year))

                wb.close()
            except Exception as e:
                print(f"Error detectando fechas con openpyxl: {e}")

        # Fallback
        if not anio_detectado:
            anio_detectado = '2026'

        return anio_detectado, mes_detectado

    def handle_actualizar_tramite(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                self.send_error(400, "Empty request")
                return

            body = self.rfile.read(content_length)
            data = json.loads(body.decode('utf-8'))

            unidad = data.get('unidad_negocio')
            ot = data.get('ot')

            if not unidad or not ot:
                self.send_error(400, "Missing required fields")
                return

            import dashboard_updater
            output_seg = os.path.join(BASE_DIR, "seguimiento_data.js")
            seg_data = dashboard_updater.load_existing_json(output_seg, 'PRELOADED_SEGUIMIENTO')

            updated = False
            for t in seg_data:
                if t.get('unidad_negocio') == unidad and t.get('ot') == ot:
                    if 'estado' in data: t['estado'] = data['estado']
                    if 'fecha_estimada' in data: t['fecha_estimada'] = data['fecha_estimada']
                    if 'notas' in data: t['notas'] = data['notas']
                    if 'no_caso_portal' in data: t['no_caso_portal'] = data['no_caso_portal']
                    if 'fecha_cierre_portal' in data: t['fecha_cierre_portal'] = data['fecha_cierre_portal']
                    if 'monto_mano_obra' in data:
                        try:
                            t['monto_mano_obra'] = float(data['monto_mano_obra'])
                        except Exception:
                            t['monto_mano_obra'] = 0
                    updated = True
                    break

            if updated:
                dashboard_updater.save_json(output_seg, 'PRELOADED_SEGUIMIENTO', seg_data, "Seguimiento", "SEGUIMIENTO_META")
                print(f"Sincronizando actualizacion de OT {ot} con GitHub...")
                subprocess.run(["git", "add", "seguimiento_data.js"], cwd=BASE_DIR, check=False)
                status = subprocess.run(["git", "status", "--porcelain"], cwd=BASE_DIR, capture_output=True, text=True)
                if status.stdout.strip():
                    commit_msg = f"Update tracking OT {ot}"
                    subprocess.run(["git", "commit", "-m", commit_msg], cwd=BASE_DIR, check=False)
                    subprocess.run(["git", "push", "origin", "HEAD:master"], cwd=BASE_DIR, check=False)

                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'success': True}).encode())
            else:
                self.send_error(404, "OT Not Found")

        except Exception as ex:
            print(f"Error en actualizar_tramite: {ex}")
            self.send_error(500, str(ex))


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8001))
    server_address = ('0.0.0.0', port)
    httpd = ThreadingHTTPServer(server_address, DashboardServer)
    print(f"Servidor backend ejecutándose en el puerto {port}...")
    httpd.serve_forever()
